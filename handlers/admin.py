import io
import os
import hashlib
import aiohttp
import openpyxl
from aiogram import Router, F, Bot
from aiogram.types import Message, BufferedInputFile, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.context import FSMContext

from database.db import (
    get_all_users, get_full_stats, add_question, get_all_questions,
    count_questions, delete_all_questions, delete_questions_by_filter,
    get_user_tariff, admin_grant_subscription, admin_revoke_subscription,
    grant_attestation, ban_user, unban_user, reset_all_subscriptions
)
from keyboards.keyboards import (
    admin_keyboard, cancel_keyboard, main_menu_keyboard,
    subject_keyboard, addq_category_keyboard, addq_topic_keyboard,
    addq_grade_keyboard, correct_answer_keyboard,
    skip_image_keyboard
)
from states import AdminStates
from config import config

router = Router()

# ══════════════════════════════════════════════
# ADMIN PANEL
# ══════════════════════════════════════════════

def is_admin(message: Message) -> bool:
    return message.from_user.id in config.ADMIN_IDS

# ══════════════════════════════════════════════
# FOYDALANUVCHILAR RO'YXATI
# ══════════════════════════════════════════════

# ══════════════════════════════════════════════
# STATISTIKA
# ══════════════════════════════════════════════

@router.message(F.text == "📊 Statistika")
async def admin_stats(message: Message):
    if message.from_user.id not in config.ADMIN_IDS:
        return
    s = await get_full_stats()
    await message.answer(
        f"📊 <b>Statistika</b>\n\n"
        f"👥 Jami foydalanuvchi: <b>{s['total_users']}</b>\n"
        f"✅ Ro'yxatdan o'tgan: <b>{s['registered']}</b>\n\n"
        f"💳 <b>To'lovlar:</b>\n"
        f"⏳ Kutayotgan: <b>{s['pending']}</b>\n"
        f"✅ Tasdiqlangan: <b>{s['confirmed_purchases']}</b>\n\n"
        f"📅 <b>Faol obunalar:</b>\n"
        f"Kunlik: <b>{s['active_daily']}</b>  |  "
        f"Oylik: <b>{s['active_monthly']}</b>\n\n"
        f"📝 <b>Testlar:</b>\n"
        f"Jami: <b>{s['total_tests']}</b>  |  "
        f"Bugun: <b>{s['today_tests']}</b>\n"
        f"📈 O'rtacha ball: <b>{s['avg_score']}%</b>\n\n"
        f"❓ <b>Savollar:</b>\n"
        f"Jami: <b>{s['total_questions']}</b>\n"
        f"📚 Ona tili: <b>{s['onatili_q']}</b>\n"
        f"📖 Adabiyot: <b>{s['adabiyot_q']}</b>\n"
        f"🎓 Attestatsiya: <b>{s['attestation_q']}</b>\n"
        f"🏅 Milliy: <b>{s['milliy_q']}</b>",
        parse_mode="HTML"
    )


# ══════════════════════════════════════════════
# FOYDALANUVCHILAR
# ══════════════════════════════════════════════

@router.message(F.text == "👥 Foydalanuvchilar")
async def admin_users(message: Message):
    if not is_admin(message): return

    users = await get_all_users()
    total = len(users)
    text  = f"👥 <b>Foydalanuvchilar: {total} ta</b>\n\n"

    for u in users[:20]:
        icon  = "✅" if u.is_registered else "👤"
        uname = f"@{u.username}" if u.username else "—"
        text += f"{icon} {u.full_name or 'Noma\'lum'} | {u.phone_number or '—'} | {uname}\n"

    if total > 20:
        text += f"\n... va yana {total - 20} ta"

    await message.answer(text, parse_mode="HTML")

# ══════════════════════════════════════════════
# EXCEL EKSPORT
# ══════════════════════════════════════════════

@router.message(F.text == "📥 Excel eksport")
async def admin_export(message: Message):
    if not is_admin(message): return
    await message.answer("⏳ Fayllar tayyorlanmoqda...")

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    # ── 1. Foydalanuvchilar ──────────────────────
    users = await get_all_users()
    wb    = openpyxl.Workbook()
    ws    = wb.active
    ws.title = "Foydalanuvchilar"

    hfill = PatternFill("solid", fgColor="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    headers = ["#", "Telegram ID", "Ism", "Username", "Telefon", "Tarif", "Ro'yxat", "Sana"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i, h)
        cell.font  = hfont
        cell.fill  = hfill
        cell.alignment = Alignment(horizontal="center")

    TARIFF_LABELS = {"daily": "Kunlik", "monthly": "Oylik", "yearly": "Yillik",
                     "attestation": "Attestatsiya", "free": "Bepul"}
    for i, u in enumerate(users, 2):
        tariff = await get_user_tariff(u.telegram_id)
        ws.append([
            i - 1,
            u.telegram_id,
            u.full_name or "",
            f"@{u.username}" if u.username else "",
            u.phone_number or "",
            TARIFF_LABELS.get(tariff["type"], "Bepul"),
            "Ha" if u.is_registered else "Yo'q",
            str(u.registered_at)[:16] if u.registered_at else "",
        ])

    col_widths = [5, 15, 25, 20, 16, 12, 10, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 2. Savollar ─────────────────────────────
    questions = await get_all_questions()
    ws2 = wb.create_sheet("Savollar")

    hfill2 = PatternFill("solid", fgColor="0F4C81")
    headers2 = ["#", "subject", "category", "subcategory", "is_attestation",
                "order_num", "question", "a", "b", "c", "d", "correct",
                "question_type", "written_parts", "keywords_1", "keywords_2"]
    for i, h in enumerate(headers2, 1):
        cell = ws2.cell(1, i, h)
        cell.font  = hfont
        cell.fill  = hfill2
        cell.alignment = Alignment(horizontal="center")

    SUBJ_COLORS = {
        "onatili": "EBF3FB", "adabiyot": "EBF7EE",
        "attestation": "FFF2CC", "milliy": "E8F5E9"
    }
    for i, q in enumerate(questions, 2):
        color = SUBJ_COLORS.get(q.subject, "FFFFFF")
        fill  = PatternFill("solid", fgColor=color)
        row = [
            i - 1,
            q.subject, q.category, q.subcategory or "",
            "TRUE" if q.is_attestation else "FALSE",
            q.order_num or "",
            q.question_text,
            q.option_a or "", q.option_b or "",
            q.option_c or "", q.option_d or "",
            q.correct_answer or "",
            q.question_type or "choice",
            q.written_parts or 1,
            q.keywords_1 or "", q.keywords_2 or "",
        ]
        for j, val in enumerate(row, 1):
            cell = ws2.cell(i, j, val)
            cell.fill = fill

    col_widths2 = [5, 12, 14, 30, 14, 10, 50, 20, 20, 20, 20, 10, 14, 14, 30, 30]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    await message.answer_document(
        document=BufferedInputFile(buf.read(), filename="eksport.xlsx"),
        caption=(
            f"📥 <b>Eksport tayyor!</b>\n\n"
            f"👥 Foydalanuvchilar: <b>{len(users)} ta</b>\n"
            f"📝 Savollar: <b>{len(questions)} ta</b>"
        ),
        parse_mode="HTML"
    )

# ══════════════════════════════════════════════
# BROADCAST
# ══════════════════════════════════════════════

@router.message(F.text == "📢 Broadcast")
async def admin_broadcast_start(message: Message, state: FSMContext):
    if not is_admin(message): return
    await message.answer(
        "📢 <b>Broadcast</b>\n\n"
        "Barcha foydalanuvchilarga yuboriladigan xabarni yozing\n"
        "(matn, rasm yoki video yuborishingiz mumkin):",
        reply_markup=cancel_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state(AdminStates.broadcast_message)

@router.message(AdminStates.broadcast_message, F.text == "❌ Bekor qilish")
async def broadcast_cancel(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("❌ Bekor qilindi.", reply_markup=admin_keyboard())

@router.message(AdminStates.broadcast_message)
async def admin_broadcast_send(message: Message, state: FSMContext, bot: Bot):
    await state.clear()
    users   = await get_all_users()
    success = 0
    failed  = 0

    await message.answer(f"⏳ {len(users)} ta foydalanuvchiga yuborilmoqda...")

    for user in users:
        try:
            if message.photo:
                await bot.send_photo(
                    chat_id=user.telegram_id,
                    photo=message.photo[-1].file_id,
                    caption=message.caption or ""
                )
            elif message.video:
                await bot.send_video(
                    chat_id=user.telegram_id,
                    video=message.video.file_id,
                    caption=message.caption or ""
                )
            else:
                await bot.send_message(
                    chat_id=user.telegram_id,
                    text=message.text or ""
                )
            success += 1
        except Exception:
            failed += 1

    await message.answer(
        f"📢 <b>Broadcast tugadi!</b>\n\n"
        f"✅ Yuborildi: <b>{success}</b>\n"
        f"❌ Xato:     <b>{failed}</b>",
        reply_markup=admin_keyboard(),
        parse_mode="HTML"
    )

# ══════════════════════════════════════════════
# SAVOL QO'SHISH
# ══════════════════════════════════════════════

@router.message(F.text == "➕ Savol qo'shish")
async def add_question_start(message: Message, state: FSMContext):
    if not is_admin(message): return
    await message.answer(
        "➕ <b>Yangi savol qo'shish</b>\n\nFanni tanlang:",
        reply_markup=subject_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state(AdminStates.add_subject)

# Fan tanlash
@router.callback_query(F.data.startswith("addq:subject:"))
async def addq_subject(callback: CallbackQuery, state: FSMContext):
    subject = callback.data.split(":")[2]
    await state.update_data(subject=subject)
    SUBJ = {
        'onatili':     '📚 Ona tili',
        'adabiyot':    '📖 Adabiyot',
        'attestation': '🎓 Attestatsiya',
        'milliy':      '🏅 Milliy sertifikat',
    }
    if subject in ('attestation', 'milliy'):
        await state.update_data(category=subject, subcategory=None, is_attestation=True)
        await callback.message.edit_text(
            f"📁 <b>{SUBJ.get(subject)}</b>\n\nSavol turini tanlang:",
            reply_markup=addq_category_keyboard(subject),
            parse_mode="HTML"
        )
        await state.set_state(AdminStates.add_is_attest)
    else:
        await callback.message.edit_text(
            f"📁 <b>{SUBJ.get(subject)} — Kategoriya tanlang:</b>",
            reply_markup=addq_category_keyboard(subject),
            parse_mode="HTML"
        )
        await state.set_state(AdminStates.add_category)
    await callback.answer()

# Kategoriya tanlash
@router.callback_query(F.data.startswith("addq:cat:"))
async def addq_category(callback: CallbackQuery, state: FSMContext):
    category = callback.data.split(":")[2]
    await state.update_data(category=category)
    data = await state.get_data()
    subject = data.get('subject')

    if category == 'mavzu':
        await callback.message.edit_text(
            "📌 <b>Mavzuni tanlang:</b>",
            reply_markup=addq_topic_keyboard(),
            parse_mode="HTML"
        )
        await state.set_state(AdminStates.add_subcategory)

    elif category == 'sinf':
        await callback.message.edit_text(
            "🏫 <b>Sinfni tanlang:</b>",
            reply_markup=addq_grade_keyboard(),
            parse_mode="HTML"
        )
        await state.set_state(AdminStates.add_subcategory)

    else:
        # aralash, gazallar, sheriy, badiiy — to'g'ridan rasmga
        await state.update_data(is_attestation=False, subcategory=None)
        await callback.message.answer(
            "📸 Rasm yuboring yoki o'tkazib yuboring:",
            reply_markup=skip_image_keyboard()
        )
        await state.set_state(AdminStates.add_image)

    await callback.answer()

# Subcategory (mavzu yoki sinf)
@router.callback_query(F.data.startswith("addq:topic:"))
async def addq_topic(callback: CallbackQuery, state: FSMContext):
    topic = callback.data.split(":")[2]
    label = config.ONA_TILI_BOLIMLAR.get(topic, topic)
    submavzular = config.ONA_TILI_SUBMAVZULAR.get(topic, {})

    if submavzular:
        # Sub-mavzular bor — tanlash kerak
        await state.update_data(bolim=topic, is_attestation=False)
        from keyboards.keyboards import addq_submavzu_keyboard
        await callback.message.edit_text(
            f"📌 <b>{label}</b>\n\nSub-mavzuni tanlang:",
            reply_markup=addq_submavzu_keyboard(topic),
            parse_mode="HTML"
        )
    else:
        # Sub-mavzu yo'q — to'g'ridan rasmga
        await state.update_data(subcategory=topic, is_attestation=False)
        await callback.message.answer(
            f"📌 <b>{label}</b>\n\n📸 Rasm yuboring yoki o'tkazib yuboring:",
            reply_markup=skip_image_keyboard(),
            parse_mode="HTML"
        )
        await state.set_state(AdminStates.add_image)
    await callback.answer()

# Sub-mavzu tanlash
@router.callback_query(F.data.startswith("addq:sub:"))
async def addq_sub(callback: CallbackQuery, state: FSMContext):
    sub_key = callback.data.split(":")[2]
    data = await state.get_data()
    bolim = data.get("bolim", "")
    # subcategory = bolim_submavzu formatda
    subcategory = f"{bolim}_{sub_key}" if bolim else sub_key
    await state.update_data(subcategory=subcategory, is_attestation=False)
    label = config.ONA_TILI_SUBMAVZULAR.get(bolim, {}).get(sub_key, sub_key)
    await callback.message.answer(
        f"📌 <b>{label}</b>\n\n📸 Rasm yuboring yoki o'tkazib yuboring:",
        reply_markup=skip_image_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state(AdminStates.add_image)
    await callback.answer()

@router.callback_query(F.data.startswith("addq:grade:"))
async def addq_grade(callback: CallbackQuery, state: FSMContext):
    grade      = callback.data.split(":")[2]
    grade_name = config.GRADES.get(grade, f"{grade}-sinf")
    boblar     = config.ADABIYOT_BOBLAR.get(grade, {})

    if boblar:
        # Bob tanlash oqimi
        await state.update_data(grade=grade, is_attestation=False)
        from keyboards.keyboards import addq_bob_keyboard
        await callback.message.edit_text(
            f"🏫 <b>{grade_name}</b>\n\nBobni tanlang:",
            reply_markup=addq_bob_keyboard(grade),
            parse_mode="HTML"
        )
    else:
        # Bob yo'q — to'g'ridan rasm
        await state.update_data(subcategory=grade, is_attestation=False)
        await callback.message.answer(
            f"🏫 <b>{grade_name}</b>\n\n📸 Rasm yuboring yoki o'tkazib yuboring:",
            reply_markup=skip_image_keyboard(),
            parse_mode="HTML"
        )
        await state.set_state(AdminStates.add_image)
    await callback.answer()


@router.callback_query(F.data.startswith("addq:bob:"))
async def addq_bob(callback: CallbackQuery, state: FSMContext):
    bob_key = callback.data.split(":")[2]
    data    = await state.get_data()
    grade   = data.get("grade", "")
    subcategory = f"{grade}_{bob_key}" if grade else bob_key
    grade_name  = config.GRADES.get(grade, f"{grade}-sinf")
    bob_name    = config.ADABIYOT_BOBLAR.get(grade, {}).get(bob_key, f"{bob_key}-bob")

    await state.update_data(subcategory=subcategory, is_attestation=False)
    await callback.message.answer(
        f"📖 <b>{grade_name} — {bob_name}</b>\n\n📸 Rasm yuboring yoki o'tkazib yuboring:",
        reply_markup=skip_image_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state(AdminStates.add_image)
    await callback.answer()


# Order num (attestation uchun)
@router.message(AdminStates.add_order_num)
async def addq_order_num(message: Message, state: FSMContext):
    try:
        order_num = int(message.text.strip())
        await state.update_data(order_num=order_num)
    except ValueError:
        await message.answer("❌ Raqam kiriting!")
        return

    await message.answer(
        "📸 Rasm yuboring yoki o'tkazib yuboring:",
        reply_markup=skip_image_keyboard()
    )
    await state.set_state(AdminStates.add_image)

# Rasm
@router.message(AdminStates.add_image, F.photo)
async def addq_image(message: Message, state: FSMContext):
    await state.update_data(image_file_id=message.photo[-1].file_id)
    await message.answer(
        "✅ Rasm saqlandi.\n\n📝 Savol matnini yozing:",
        reply_markup=cancel_keyboard()
    )
    await state.set_state(AdminStates.add_text)

@router.message(AdminStates.add_image, F.text == "⏭ Rasmisiz davom etish")
async def addq_skip_image(message: Message, state: FSMContext):
    await state.update_data(image_file_id=None)
    await message.answer("📝 Savol matnini yozing:", reply_markup=cancel_keyboard())
    await state.set_state(AdminStates.add_text)

# Savol matni
@router.message(AdminStates.add_text)
async def addq_text(message: Message, state: FSMContext):
    if message.text == "❌ Bekor qilish":
        await state.clear()
        await message.answer("❌ Bekor qilindi.", reply_markup=admin_keyboard())
        return
    await state.update_data(question_text=message.text)
    await message.answer("🅰 A variantini yozing:")
    await state.set_state(AdminStates.add_a)

# Variantlar
@router.message(AdminStates.add_a)
async def addq_a(message: Message, state: FSMContext):
    await state.update_data(option_a=message.text)
    await message.answer("🅱 B variantini yozing:")
    await state.set_state(AdminStates.add_b)

@router.message(AdminStates.add_b)
async def addq_b(message: Message, state: FSMContext):
    await state.update_data(option_b=message.text)
    await message.answer("🅲 C variantini yozing:")
    await state.set_state(AdminStates.add_c)

@router.message(AdminStates.add_c)
async def addq_c(message: Message, state: FSMContext):
    await state.update_data(option_c=message.text)
    await message.answer("🅳 D variantini yozing:")
    await state.set_state(AdminStates.add_d)

@router.message(AdminStates.add_d)
async def addq_d(message: Message, state: FSMContext):
    await state.update_data(option_d=message.text)
    await message.answer(
        "✅ <b>To'g'ri javobni tanlang:</b>",
        reply_markup=correct_answer_keyboard(),
        parse_mode="HTML"
    )
    await state.set_state(AdminStates.add_correct)

# Savol turi tanlash (attestation/milliy uchun)
# Attestatsiya bo'lim tanlash
@router.callback_query(F.data.startswith("addq:attest_bolim:"))
async def addq_attest_bolim(callback: CallbackQuery, state: FSMContext):
    bolim_num = int(callback.data.split(":")[2])
    if bolim_num == 0:
        subcategory = None  # Umumiy
        bolim_label = "Barcha bo'limlar (umumiy)"
    else:
        subcategory = f"bolim_{bolim_num}"
        bolim_label = f"{bolim_num}-bo'lim"

    await state.update_data(subcategory=subcategory, bolim_label=bolim_label)

    from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
    await callback.message.edit_text(
        f"🎓 <b>Attestatsiya — {bolim_label}</b>\n\nSavol turini tanlang:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔘 Variantli (A/B/C/D)", callback_data="addq:qtype:choice")],
            [InlineKeyboardButton(text="❌ Bekor", callback_data="addq:cancel")],
        ]),
        parse_mode="HTML"
    )
    await state.set_state(AdminStates.add_is_attest)
    await callback.answer()


@router.callback_query(F.data.startswith("addq:qtype:"))
async def addq_qtype(callback: CallbackQuery, state: FSMContext):
    qtype   = callback.data.split(":")[2]
    data    = await state.get_data()
    subject = data.get('subject', 'onatili')
    LABELS  = {'attestation': '🎓 Attestatsiya', 'milliy': '🏅 Milliy sertifikat'}
    label   = LABELS.get(subject, subject)

    if qtype == 'choice':
        await state.update_data(question_type='choice', written_parts=1)
        data = await state.get_data()
        subject = data.get('subject', 'onatili')
        subcategory = data.get('subcategory')
        cnt = await count_questions(subject=subject, category=subject,
                                    subcategory=subcategory, is_attestation=True)
        bolim_label = data.get('bolim_label', '')
        bolim_info  = f" — {bolim_label}" if bolim_label else ""
        await callback.message.edit_text(
            f"{label}{bolim_info} — Variantli savol\n\nTartib raqamini yozing (<code>{cnt + 1}</code>):",
            parse_mode="HTML"
        )
        await state.set_state(AdminStates.add_order_num)
    elif qtype in ('written1', 'written2'):
        parts = 1 if qtype == 'written1' else 2
        await state.update_data(question_type='written', written_parts=parts)
        data = await state.get_data()
        subject = data.get('subject', 'onatili')
        subcategory = data.get('subcategory')
        cnt = await count_questions(subject=subject, category=subject,
                                    subcategory=subcategory, is_attestation=True)
        await callback.message.edit_text(
            f"{label} — Yozma savol ({parts} qism)\n\nTartib raqamini yozing (<code>{cnt + 1}</code>):",
            parse_mode="HTML"
        )
        await state.set_state(AdminStates.add_order_num)
    await callback.answer()

# To'g'ri javob
@router.callback_query(F.data.startswith("addq:correct:"))
async def addq_correct(callback: CallbackQuery, state: FSMContext):
    correct = callback.data.split(":")[2]
    data    = await state.get_data()
    await state.clear()

    subj = data['subject']
    cat  = data.get('category', subj)  # attestation/milliy uchun category=subject
    subcat = data.get('subcategory')

    await add_question(
        subject        = subj,
        category       = cat,
        question_text  = data['question_text'],
        option_a       = data.get('option_a'),
        option_b       = data.get('option_b'),
        option_c       = data.get('option_c'),
        option_d       = data.get('option_d'),
        correct_answer = correct,
        subcategory    = subcat,
        difficulty     = None,
        is_attestation = data.get('is_attestation', False),
        order_num      = data.get('order_num'),
        image_file_id  = data.get('image_file_id'),
        question_type  = data.get('question_type', 'choice'),
        written_parts  = data.get('written_parts', 1),
    )

    SUBJ = {'onatili': '📚 Ona tili', 'adabiyot': '📖 Adabiyot',
             'attestation': '🎓 Attestatsiya', 'milliy': '🏅 Milliy sertifikat'}
    bolim_label = data.get('bolim_label', '')
    bolim_info  = f" — {bolim_label}" if bolim_label else (f" — {subcat}" if subcat else "")
    await callback.message.edit_text(
        f"✅ <b>Savol qo'shildi!</b>\n\n"
        f"📚 {SUBJ.get(subj, subj)}{bolim_info}\n"
        f"✅ To'g'ri javob: <b>{correct}</b>",
        parse_mode="HTML"
    )
    await callback.message.answer(
        "Yana savol qo'shish uchun '➕ Savol qo'shish' tugmasini bosing.",
        reply_markup=admin_keyboard()
    )
    await callback.answer()

# Bekor qilish callback
@router.callback_query(F.data == "addq:cancel")
async def addq_cancel(callback: CallbackQuery, state: FSMContext):
    await state.clear()
    try:
        await callback.message.edit_text("❌ Bekor qilindi.")
    except Exception:
        pass
    await callback.bot.send_message(
        chat_id      = callback.from_user.id,
        text         = "Admin panel:",
        reply_markup = admin_keyboard()
    )
    await callback.answer()

# ══════════════════════════════════════════════
# BARCHA SAVOLLARNI O'CHIRISH
# ══════════════════════════════════════════════

# ══════════════════════════════════════════════
# BARCHA TARIFLARNI NOLLASH
# ══════════════════════════════════════════════

@router.message(F.text == "♻️ Tariflarni nollash")
async def reset_tariffs_confirm(message: Message):
    if not is_admin(message): return
    await message.answer(
        "⚠️ <b>Diqqat!</b>\n\n"
        "Barcha foydalanuvchilarning <b>kunlik, oylik va attestatsiya</b> huquqlari "
        "to'liq bekor qilinadi.\n"
        "Hammasi qaytadan sotib olishi kerak bo'ladi.\n\n"
        "Tasdiqlaysizmi?",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text="♻️ Ha, nollash", callback_data="admin:reset_tariffs"),
            InlineKeyboardButton(text="❌ Yo'q",         callback_data="admin:cancel_delete"),
        ]]),
        parse_mode="HTML"
    )

@router.callback_query(F.data == "admin:reset_tariffs")
async def reset_tariffs_execute(callback: CallbackQuery):
    count = await reset_all_subscriptions()
    await callback.message.edit_text(
        f"♻️ <b>{count} ta</b> tarif nollandi!\n\n"
        f"Barcha foydalanuvchilar (kunlik, oylik va attestatsiya) endi qayta sotib olishlari kerak.",
        parse_mode="HTML"
    )
    await callback.message.answer("Admin panel:", reply_markup=admin_keyboard())
    await callback.answer("✅ Tariflar nollandi!")


@router.message(F.text == "🗑 Savollarni o'chirish")
async def delete_questions_confirm(message: Message):
    if not is_admin(message): return
    cnt = await count_questions()
    await message.answer(
        f"⚠️ <b>Diqqat!</b>\n\n"
        f"Hozir bazada <b>{cnt} ta</b> savol bor.\n"
        f"Barchasini o'chirishni tasdiqlaysizmi?",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="🗑 Ha, o'chirish", callback_data="admin:delete_all_q"),
                InlineKeyboardButton(text="❌ Yo'q", callback_data="admin:cancel_delete"),
            ]
        ]),
        parse_mode="HTML"
    )


@router.callback_query(F.data == "admin:delete_all_q")
async def delete_questions_execute(callback: CallbackQuery):
    deleted = await delete_all_questions()
    await callback.message.edit_text(
        f"🗑 <b>{deleted} ta savol o'chirildi!</b>",
        reply_markup=admin_keyboard(),
        parse_mode="HTML"
    )
    await callback.answer()

@router.callback_query(F.data == "admin:cancel_delete")
async def delete_questions_cancel(callback: CallbackQuery):
    await callback.message.edit_text("❌ Bekor qilindi.")
    await callback.answer()

# ══════════════════════════════════════════════
# BO'LIM BOSHQARISH (O'CHIRISH / QO'SHISH)
# ══════════════════════════════════════════════

@router.message(F.text == "🗂 Bo'lim o'chirish")
async def section_manage_start(message: Message):
    if not is_admin(message): return
    await message.answer(
        "🗂 <b>Qaysi fan bo'limini boshqarish kerak?</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📚 Ona tili",          callback_data="smng:onatili")],
            [InlineKeyboardButton(text="📖 Adabiyot",          callback_data="smng:adabiyot")],
            [InlineKeyboardButton(text="🎓 Attestatsiya",      callback_data="smng:attestation")],
            [InlineKeyboardButton(text="🏅 Milliy sertifikat", callback_data="smng:milliy")],
            [InlineKeyboardButton(text="❌ Bekor",             callback_data="smng:cancel")],
        ]),
        parse_mode="HTML"
    )

# ── Fan tanlandi ──
@router.callback_query(F.data.startswith("smng:"))
async def section_manage_subject(callback: CallbackQuery):
    subject = callback.data.split(":")[1]
    if subject == "cancel":
        await callback.message.edit_text("❌ Bekor qilindi.")
        await callback.answer()
        return

    LABELS = {'onatili': '📚 Ona tili', 'adabiyot': '📖 Adabiyot',
              'attestation': '🎓 Attestatsiya', 'milliy': '🏅 Milliy sertifikat'}

    if subject == 'adabiyot':
        # Sinflarni ko'rsatish
        btns = []
        for grade, label in config.GRADES.items():
            # grade_bob formatdagi savollar: 10_1, 10_2 ... prefix = "10_"
            cnt = await count_questions(subject='adabiyot', category='sinf',
                                        subcategory_prefix=f"{grade}_")
            # + sinf aralash (subcategory=grade)
            cnt += await count_questions(subject='adabiyot', category='sinf', subcategory=grade)
            btns.append([InlineKeyboardButton(
                text=f"{label} ({cnt} savol)",
                callback_data=f"smng_grade:{grade}"
            )])
        # Adabiyot kategoriyalari
        for cat in ['aralash', 'gazallar', 'sheriy', 'badiiy']:
            cnt = await count_questions(subject='adabiyot', category=cat)
            btns.append([InlineKeyboardButton(
                text=f"📂 {cat.capitalize()} ({cnt} savol)",
                callback_data=f"smng_cat:adabiyot:{cat}"
            )])
        btns.append([InlineKeyboardButton(text="⬅️ Orqaga", callback_data="smng:cancel")])
        await callback.message.edit_text(
            f"📖 <b>Adabiyot — bo'limni tanlang:</b>",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=btns),
            parse_mode="HTML"
        )

    elif subject == 'onatili':
        # Ona tili bo'limlari
        btns = []
        for key, label in config.ONA_TILI_BOLIMLAR.items():
            # key va key_submavzu formatdagi barchasi: prefix = "key"
            cnt = await count_questions(subject='onatili', category='mavzu',
                                        subcategory_prefix=key)
            btns.append([InlineKeyboardButton(
                text=f"{label} ({cnt} savol)",
                callback_data=f"smng_onatili:{key}"
            )])
        for cat in ['aralash']:
            cnt = await count_questions(subject='onatili', category=cat)
            btns.append([InlineKeyboardButton(
                text=f"📂 Aralash ({cnt} savol)",
                callback_data=f"smng_cat:onatili:aralash"
            )])
        btns.append([InlineKeyboardButton(text="⬅️ Orqaga", callback_data="smng:cancel")])
        await callback.message.edit_text(
            "📚 <b>Ona tili — bo'limni tanlang:</b>",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=btns),
            parse_mode="HTML"
        )

    else:
        # Attestatsiya / Milliy — to'g'ridan tasdiqlash
        cnt = await count_questions(subject=subject, is_attestation=True)
        await callback.message.edit_text(
            f"⚠️ <b>{LABELS.get(subject)}</b>\n\n"
            f"<b>{cnt} ta</b> savol bor.\n"
            f"Barchasini o'chirishni tasdiqlaysizmi?",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"smng_del:{subject}:all:all"),
                InlineKeyboardButton(text="❌ Yo'q",      callback_data="smng:cancel"),
            ]]),
            parse_mode="HTML"
        )
    await callback.answer()

# ── Adabiyot sinf tanlandi ──
@router.callback_query(F.data.startswith("smng_grade:"))
async def section_manage_grade(callback: CallbackQuery):
    grade = callback.data.split(":")[1]
    label = config.GRADES.get(grade, grade)
    boblar = config.ADABIYOT_BOBLAR.get(grade, {})

    btns = []
    for bob_key, bob_label in boblar.items():
        sub = f"{grade}_{bob_key}"
        cnt = await count_questions(subject='adabiyot', category='sinf', subcategory=sub)
        btns.append([InlineKeyboardButton(
            text=f"{bob_label} ({cnt} savol)",
            callback_data=f"smng_bob:{grade}:{bob_key}"
        )])
    # Sinf umumiy (barcha boblar)
    cnt_all = await count_questions(subject='adabiyot', category='sinf',
                                    subcategory_prefix=f"{grade}_")
    cnt_all += await count_questions(subject='adabiyot', category='sinf', subcategory=grade)
    btns.append([InlineKeyboardButton(
        text=f"🗑 Butun {label}ni o'chirish ({cnt_all} savol)",
        callback_data=f"smng_del:adabiyot:sinf_all:{grade}"
    )])
    btns.append([InlineKeyboardButton(text="⬅️ Orqaga", callback_data="smng:adabiyot")])

    await callback.message.edit_text(
        f"📖 <b>{label} — bobni tanlang:</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=btns),
        parse_mode="HTML"
    )
    await callback.answer()

# ── Bob tanlandi ──
@router.callback_query(F.data.startswith("smng_bob:"))
async def section_manage_bob(callback: CallbackQuery):
    _, grade, bob_key = callback.data.split(":")
    sub   = f"{grade}_{bob_key}"
    label = f"{config.GRADES.get(grade, grade)} {config.ADABIYOT_BOBLAR.get(grade, {}).get(bob_key, bob_key)}"
    cnt   = await count_questions(subject='adabiyot', category='sinf', subcategory=sub)

    await callback.message.edit_text(
        f"📖 <b>{label}</b>\n\n"
        f"<b>{cnt} ta</b> savol bor.\n\n"
        f"Nima qilmoqchisiz?",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=f"🗑 O'chirish ({cnt} ta)",
                                  callback_data=f"smng_del:adabiyot:sinf:{sub}")],
            [InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"smng_grade:{grade}")],
        ]),
        parse_mode="HTML"
    )
    await callback.answer()

# ── Ona tili bo'lim tanlandi ──
@router.callback_query(F.data.startswith("smng_onatili:"))
async def section_manage_onatili(callback: CallbackQuery):
    bolim = callback.data.split(":")[1]
    label = config.ONA_TILI_BOLIMLAR.get(bolim, bolim)
    submavzular = config.ONA_TILI_SUBMAVZULAR.get(bolim, {})

    btns = []
    if submavzular:
        for sub_key, sub_label in submavzular.items():
            sub = f"{bolim}_{sub_key}"
            cnt = await count_questions(subject='onatili', category='mavzu', subcategory=sub)
            btns.append([InlineKeyboardButton(
                text=f"{sub_label} ({cnt} savol)",
                callback_data=f"smng_sub:{bolim}:{sub_key}"
            )])
    # Umumiy bo'lim (barcha submavzular bilan)
    cnt_all = await count_questions(subject='onatili', category='mavzu',
                                    subcategory_prefix=bolim)
    btns.append([InlineKeyboardButton(
        text=f"🗑 Butun bo'limni o'chirish ({cnt_all} savol)",
        callback_data=f"smng_del:onatili:mavzu_all:{bolim}"
    )])
    btns.append([InlineKeyboardButton(text="⬅️ Orqaga", callback_data="smng:onatili")])

    await callback.message.edit_text(
        f"📚 <b>{label}</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=btns),
        parse_mode="HTML"
    )
    await callback.answer()

# ── Ona tili submavzu tanlandi ──
@router.callback_query(F.data.startswith("smng_sub:"))
async def section_manage_sub(callback: CallbackQuery):
    _, bolim, sub_key = callback.data.split(":")
    sub   = f"{bolim}_{sub_key}"
    label = config.ONA_TILI_SUBMAVZULAR.get(bolim, {}).get(sub_key, sub_key)
    cnt   = await count_questions(subject='onatili', category='mavzu', subcategory=sub)

    await callback.message.edit_text(
        f"📚 <b>{label}</b>\n\n"
        f"<b>{cnt} ta</b> savol bor.\n\n"
        f"Nima qilmoqchisiz?",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=f"🗑 O'chirish ({cnt} ta)",
                                  callback_data=f"smng_del:onatili:mavzu:{sub}")],
            [InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"smng_onatili:{bolim}")],
        ]),
        parse_mode="HTML"
    )
    await callback.answer()

# ── Kategoriya tanlandi (aralash, gazallar va h.k.) ──
@router.callback_query(F.data.startswith("smng_cat:"))
async def section_manage_cat(callback: CallbackQuery):
    _, subject, category = callback.data.split(":")
    cnt = await count_questions(subject=subject, category=category)
    SUBJ = {'onatili': '📚 Ona tili', 'adabiyot': '📖 Adabiyot'}

    await callback.message.edit_text(
        f"{SUBJ.get(subject)} — <b>{category.capitalize()}</b>\n\n"
        f"<b>{cnt} ta</b> savol bor.\n\n"
        f"Nima qilmoqchisiz?",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text=f"🗑 O'chirish ({cnt} ta)",
                                  callback_data=f"smng_del:{subject}:{category}:all")],
            [InlineKeyboardButton(text="⬅️ Orqaga", callback_data=f"smng:{subject}")],
        ]),
        parse_mode="HTML"
    )
    await callback.answer()

# ── O'chirish tasdiqlash ──
@router.callback_query(F.data.startswith("smng_del:"))
async def section_delete_confirm(callback: CallbackQuery):
    parts    = callback.data.split(":")
    subject  = parts[1]
    category = parts[2]   # sinf_all | mavzu_all | sinf | mavzu | aralash | all
    sub      = parts[3] if len(parts) > 3 else 'all'

    # Savollar sonini hisoblash
    if category == 'sinf_all':
        # Butun sinf: grade_ prefix + grade subcategory
        cnt  = await count_questions(subject=subject, category='sinf', subcategory_prefix=f"{sub}_")
        cnt += await count_questions(subject=subject, category='sinf', subcategory=sub)
        cat_label = f"{config.GRADES.get(sub, sub)} (barcha boblar)"
    elif category == 'mavzu_all':
        # Butun ona tili bo'limi: bolim prefix
        cnt = await count_questions(subject=subject, category='mavzu', subcategory_prefix=sub)
        cat_label = f"{config.ONA_TILI_BOLIMLAR.get(sub, sub)} (barcha submavzular)"
    elif sub == 'all' or not sub:
        cnt = await count_questions(subject=subject, category=category if category != 'all' else None)
        cat_label = category
    else:
        cnt = await count_questions(subject=subject, category=category, subcategory=sub)
        cat_label = f"{category} › {sub}"

    await callback.message.edit_text(
        f"⚠️ <b>Tasdiqlang!</b>\n\n"
        f"📂 <b>{cat_label}</b>\n"
        f"<b>{cnt} ta</b> savol o'chiriladi!",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
            InlineKeyboardButton(text=f"🗑 Ha, {cnt} ta o'chirish",
                                 callback_data=f"smng_del_ok:{subject}:{category}:{sub}"),
            InlineKeyboardButton(text="❌ Bekor", callback_data="smng:cancel"),
        ]]),
        parse_mode="HTML"
    )
    await callback.answer()

# ── O'chirishni bajarish ──
@router.callback_query(F.data.startswith("smng_del_ok:"))
async def section_delete_execute(callback: CallbackQuery):
    parts    = callback.data.split(":")
    subject  = parts[1]
    category = parts[2]
    sub      = parts[3] if len(parts) > 3 else 'all'

    # Faqat savollar o'chiriladi — config (bo'lim tuzilmasi) o'zgarmaydi
    if category == 'sinf_all':
        d1 = await delete_questions_by_filter(subject=subject, category='sinf',
                                               subcategory_prefix=f"{sub}_")
        d2 = await delete_questions_by_filter(subject=subject, category='sinf', subcategory=sub)
        deleted = d1 + d2
    elif category == 'mavzu_all':
        deleted = await delete_questions_by_filter(subject=subject, category='mavzu',
                                                    subcategory_prefix=sub)
    elif sub == 'all':
        deleted = await delete_questions_by_filter(
            subject=subject,
            category=category if category != 'all' else None
        )
    else:
        deleted = await delete_questions_by_filter(
            subject=subject, category=category, subcategory=sub
        )

    await callback.message.edit_text(
        f"✅ <b>{deleted} ta</b> savol o'chirildi!\n\n"
        f"📂 Bo'lim tuzilmasi o'zgarmadi — yangi savollar qo'shishingiz mumkin.",
        reply_markup=admin_keyboard(),
        parse_mode="HTML"
    )
    await callback.answer()

# ══════════════════════════════════════════════
# BO'LIM QO'SHISH
# ══════════════════════════════════════════════

@router.message(F.text == "➕ Bo'lim qo'shish")
async def add_section_start(message: Message):
    if not is_admin(message): return
    await message.answer(
        "➕ <b>Qaysi fanga bo'lim qo'shmoqchisiz?</b>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📖 Adabiyot — yangi bob",   callback_data="addsec:adabiyot")],
            [InlineKeyboardButton(text="📚 Ona tili — yangi mavzu", callback_data="addsec:onatili")],
            [InlineKeyboardButton(text="❌ Bekor",                  callback_data="addsec:cancel")],
        ]),
        parse_mode="HTML"
    )

@router.callback_query(F.data.startswith("addsec:"))
async def add_section_choose(callback: CallbackQuery, state: FSMContext):
    subject = callback.data.split(":")[1]
    if subject == "cancel":
        await callback.message.edit_text("❌ Bekor qilindi.")
        await callback.answer()
        return

    if subject == "adabiyot":
        # Sinf tanlash
        btns = [[InlineKeyboardButton(text=label, callback_data=f"addsec_grade:{grade}")]
                for grade, label in config.GRADES.items()]
        btns.append([InlineKeyboardButton(text="❌ Bekor", callback_data="addsec:cancel")])
        await callback.message.edit_text(
            "📖 <b>Qaysi sinfga bob qo'shmoqchisiz?</b>",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=btns),
            parse_mode="HTML"
        )
    elif subject == "onatili":
        await callback.message.edit_text(
            "📚 <b>Yangi mavzu nomini yozing:</b>\n\n"
            "Format: <code>kalit_nomi|Ko'rinish nomi</code>\n"
            "Masalan: <code>leksikologiya_yangi|Yangi leksikologiya bo'limi</code>",
            parse_mode="HTML"
        )
        await state.update_data(addsec_subject="onatili")
        await state.set_state(AdminStates.add_section)
    await callback.answer()

@router.callback_query(F.data.startswith("addsec_grade:"))
async def add_section_grade(callback: CallbackQuery, state: FSMContext):
    grade = callback.data.split(":")[1]
    label = config.GRADES.get(grade, grade)
    existing = config.ADABIYOT_BOBLAR.get(grade, {})
    next_bob = str(len(existing) + 1)

    await callback.message.edit_text(
        f"📖 <b>{label}</b>\n\n"
        f"Mavjud boblar: <b>{len(existing)} ta</b>\n"
        f"Yangi bob raqami: <b>{next_bob}</b>\n\n"
        f"Bob nomini yozing (masalan: <code>5-bob</code>):",
        parse_mode="HTML"
    )
    await state.update_data(addsec_subject="adabiyot", addsec_grade=grade, addsec_bob_num=next_bob)
    await state.set_state(AdminStates.add_section)
    await callback.answer()

@router.message(AdminStates.add_section)
async def add_section_save(message: Message, state: FSMContext):
    data    = await state.get_data()
    subject = data.get("addsec_subject")
    await state.clear()

    if subject == "adabiyot":
        grade   = data.get("addsec_grade")
        bob_num = data.get("addsec_bob_num")
        bob_name = message.text.strip()

        # config ga qo'shish (runtime)
        if grade not in config.ADABIYOT_BOBLAR:
            config.ADABIYOT_BOBLAR[grade] = {}
        config.ADABIYOT_BOBLAR[grade][bob_num] = bob_name

        await message.answer(
            f"✅ <b>Qo'shildi!</b>\n\n"
            f"📖 {config.GRADES.get(grade)} — <b>{bob_name}</b>\n\n"
            f"⚠️ Bu o'zgarish bot qayta ishga tushguncha saqlanadi.\n"
            f"Doimiy saqlash uchun <code>config.py</code> ga qo'shing:\n"
            f"<code>\'{grade}\': {{...\'{bob_num}\': \'{bob_name}\'...}}</code>",
            reply_markup=admin_keyboard(),
            parse_mode="HTML"
        )

    elif subject == "onatili":
        text = message.text.strip()
        if "|" not in text:
            await message.answer(
                "❌ Format noto'g'ri!\n"
                "To'g'ri format: <code>kalit_nomi|Ko'rinish nomi</code>",
                parse_mode="HTML"
            )
            return
        key, label = text.split("|", 1)
        key   = key.strip()
        label = label.strip()
        config.ONA_TILI_BOLIMLAR[key] = label

        await message.answer(
            f"✅ <b>Qo'shildi!</b>\n\n"
            f"📚 Kalit: <code>{key}</code>\n"
            f"Ko'rinish: <b>{label}</b>\n\n"
            f"⚠️ Doimiy saqlash uchun <code>config.py</code> ga qo'shing:\n"
            f"<code>ONA_TILI_BOLIMLAR[\'{key}\'] = \'{label}\'</code>",
            reply_markup=admin_keyboard(),
            parse_mode="HTML"
        )

# ══════════════════════════════════════════════
# A'ZOLAR KO'RISH VA TARIF BOSHQARISH
# ══════════════════════════════════════════════

@router.message(F.text == "👥 A'zolar")
async def members_list(message: Message):
    if not is_admin(message): return
    users      = await get_all_users()
    total      = len(users)
    registered = sum(1 for u in users if u.is_registered)
    text = (
        f"👥 <b>A'zolar: {total} ta</b>\n"
        f"✅ Ro'yxatdan o'tgan: <b>{registered}</b>\n"
        f"👤 O'tmagan: <b>{total - registered}</b>\n\n"
        f"Batafsil ko'rish uchun:"
    )
    await message.answer(
        text,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="📋 Ro'yxat", callback_data="members:list:0")],
            [InlineKeyboardButton(text="🔍 Qidirish (ID yoki @username)", callback_data="members:search")],
        ]),
        parse_mode="HTML"
    )

@router.callback_query(F.data.startswith("members:list:"))
async def members_list_page(callback: CallbackQuery):
    offset = int(callback.data.split(":")[2])
    users  = await get_all_users()
    total  = len(users)
    page   = users[offset:offset + 25]
    TICONS = {"daily": "📅", "monthly": "📆", "yearly": "🗓", "attestation": "🎓", "free": "🆓"}
    text   = f"👥 <b>A'zolar {offset+1}–{min(offset+25, total)} / {total}</b>\n\n"
    for u in page:
        tariff = await get_user_tariff(u.telegram_id)
        icon   = TICONS.get(tariff["type"], "🆓")
        uname  = f"@{u.username}" if u.username else "—"
        phone  = u.phone_number or "—"
        date   = str(u.registered_at)[:10] if u.registered_at else "—"
        text  += f"{icon} <code>{u.telegram_id}</code> <b>{u.full_name or 'Nomsiz'}</b>\n"
        text  += f"   📱{phone} | {uname} | {date}\n"
    nav = []
    if offset > 0:
        nav.append(InlineKeyboardButton(text="⬅️", callback_data=f"members:list:{offset-25}"))
    if offset + 25 < total:
        nav.append(InlineKeyboardButton(text="➡️", callback_data=f"members:list:{offset+25}"))
    btns = []
    if nav: btns.append(nav)
    btns.append([InlineKeyboardButton(text="❌ Yopish", callback_data="members:close")])
    if len(text) > 4000:
        text = text[:4000] + "\n..."
    try:
        await callback.message.edit_text(text, reply_markup=InlineKeyboardMarkup(inline_keyboard=btns), parse_mode="HTML")
    except Exception:
        await callback.message.answer(text, reply_markup=InlineKeyboardMarkup(inline_keyboard=btns), parse_mode="HTML")
    await callback.answer()

@router.callback_query(F.data == "members:search")
async def members_search_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("🔍 Telegram ID yoki @username yozing:")
    await state.set_state(AdminStates.search_user)
    await callback.answer()

@router.message(AdminStates.search_user)
async def members_search_result(message: Message, state: FSMContext):
    await state.clear()
    query = message.text.strip().lstrip("@")
    users = await get_all_users()
    found = None
    for u in users:
        if str(u.telegram_id) == query or (u.username and u.username.lower() == query.lower()):
            found = u
            break
    if not found:
        await message.answer("❌ Foydalanuvchi topilmadi.", reply_markup=admin_keyboard())
        return
    tariff  = await get_user_tariff(found.telegram_id)
    TLABELS = {"daily": "📅 Kunlik", "monthly": "📆 Oylik", "yearly": "🗓 Yillik",
               "attestation": "🎓 Attestatsiya", "free": "🆓 Bepul"}
    tlabel  = TLABELS.get(tariff["type"], "🆓 Bepul")
    expires = f" (tugaydi: {tariff['expires']})" if tariff["expires"] else ""
    text = (
        f"👤 <b>{found.full_name or 'Nomsiz'}</b>\n\n"
        f"🆔 <code>{found.telegram_id}</code>\n"
        f"📱 {found.phone_number or '—'}\n"
        f"🔗 @{found.username or '—'}\n"
        f"📅 {str(found.registered_at)[:16] if found.registered_at else '—'}\n"
        f"💳 Tarif: <b>{tlabel}{expires}</b>"
    )
    await message.answer(
        text,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="📅 Kunlik",  callback_data=f"tariff:daily:{found.telegram_id}"),
                InlineKeyboardButton(text="📆 Oylik",   callback_data=f"tariff:monthly:{found.telegram_id}"),
            ],
            [
                InlineKeyboardButton(text="🗓 Yillik",       callback_data=f"tariff:yearly:{found.telegram_id}"),
                InlineKeyboardButton(text="🎓 Attestatsiya", callback_data=f"tariff:attestation:{found.telegram_id}"),
            ],
            [InlineKeyboardButton(text="❌ Tarifni bekor qilish", callback_data=f"tariff:revoke:{found.telegram_id}")],
            [InlineKeyboardButton(text="🚫 Bloklash",  callback_data=f"ban:block:{found.telegram_id}")],
        ]),
        parse_mode="HTML"
    )

@router.callback_query(F.data.startswith("tariff:"))
async def change_tariff(callback: CallbackQuery, bot: Bot):
    parts  = callback.data.split(":")
    action = parts[1]
    tid    = int(parts[2])
    LABELS = {
        "daily": "📅 Kunlik obuna (1 kun)",
        "monthly": "📆 Oylik obuna (30 kun)",
        "yearly": "🗓 Yillik obuna (365 kun)",
        "attestation": "🎓 Attestatsiya",
        "revoke": "❌ Bekor qilish",
    }
    label = LABELS.get(action, action)
    if action == "revoke":
        await admin_revoke_subscription(tid)
        await callback.message.edit_text(f"✅ <code>{tid}</code> — tarif bekor qilindi.", parse_mode="HTML")
    elif action == "attestation":
        await grant_attestation(tid, "attestation", "admin")
        await callback.message.edit_text(f"✅ <code>{tid}</code> — Attestatsiya berildi.", parse_mode="HTML")
        try:
            await bot.send_message(tid, "🎓 <b>Attestatsiya huquqi berildi!</b>\n\nTestni boshlashingiz mumkin.", parse_mode="HTML")
        except Exception:
            pass
    else:
        await admin_grant_subscription(tid, action)
        await callback.message.edit_text(f"✅ <code>{tid}</code> — {label} berildi.", parse_mode="HTML")
        try:
            await bot.send_message(tid, f"🎉 <b>Tarifingiz yangilandi!</b>\n\n{label} faollashtirildi.", parse_mode="HTML")
        except Exception:
            pass
    await callback.answer("✅ Tarif yangilandi!")

@router.callback_query(F.data.startswith("ban:"))
async def ban_unban_user(callback: CallbackQuery, bot: Bot):
    parts  = callback.data.split(":")
    action = parts[1]   # block | unblock
    tid    = int(parts[2])

    if action == "block":
        await ban_user(tid)
        # Foydalanuvchiga xabar
        try:
            await bot.send_message(
                tid,
                "🚫 <b>Siz botdan bloklangansiz.</b>\n\nBog'lanish uchun adminga murojaat qiling.",
                parse_mode="HTML"
            )
        except Exception:
            pass
        await callback.message.edit_text(
            f"🚫 <code>{tid}</code> — bloklandi.\n\n"
            f"<a href='tg://user?id={tid}'>Foydalanuvchiga o'tish</a>\n\n"
            f"Blokni ochish uchun:",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[[
                InlineKeyboardButton(text="✅ Blokni ochish", callback_data=f"ban:unblock:{tid}")
            ]]),
            parse_mode="HTML"
        )
        await callback.answer("🚫 Bloklandi!")

    elif action == "unblock":
        await unban_user(tid)
        try:
            await bot.send_message(
                tid,
                "✅ <b>Bloklash olib tashlandi!</b>\n\nBotdan foydalanishingiz mumkin.",
                parse_mode="HTML"
            )
        except Exception:
            pass
        await callback.message.edit_text(
            f"✅ <code>{tid}</code> — blok olib tashlandi.",
            parse_mode="HTML"
        )
        await callback.answer("✅ Blok olib tashlandi!")


@router.callback_query(F.data == "members:close")
async def members_close(callback: CallbackQuery):
    await callback.message.delete()
    await callback.answer()

# ══════════════════════════════════════════════
# YECHIM LINKI
# ══════════════════════════════════════════════

@router.message(F.text == "🔗 Yechim linki")
async def solution_url_info(message: Message):
    if not is_admin(message): return
    url = config.SOLUTION_URL
    if url:
        await message.answer(
            f"🔗 <b>Hozirgi yechim linki:</b>\n<code>{url}</code>\n\n"
            f"O'zgartirish uchun .env faylida <code>SOLUTION_URL</code> ni yangilang.",
            parse_mode="HTML"
        )
    else:
        await message.answer(
            "⚠️ <b>SOLUTION_URL</b> .env faylida yo'q!\n\n"
            "Qo'shish uchun:\n<code>SOLUTION_URL=https://t.me/kanal_nomi</code>",
            parse_mode="HTML"
        )

# ══════════════════════════════════════════════
# EXCEL DAN KO'P SAVOL YUKLASH
# ══════════════════════════════════════════════

@router.message(F.text == "📤 Excel import")
async def excel_import_start(message: Message):
    if not is_admin(message): return

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Savollar"

    headers = ["subject", "category", "subcategory",
               "is_attestation", "order_num", "question",
               "a", "b", "c", "d", "correct",
               "question_type", "written_parts", "keywords_1", "keywords_2"]

    hfill   = PatternFill("solid", fgColor="1F4E79")
    hfill_w = PatternFill("solid", fgColor="7B2D8B")
    hfont   = Font(bold=True, color="FFFFFF", size=11)

    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i)
        cell.value = h
        cell.font  = hfont
        cell.fill  = hfill_w if i >= 12 else hfill
        cell.alignment = Alignment(horizontal='center')

    widths = [12, 14, 30, 14, 10, 50, 20, 20, 20, 20, 10, 14, 14, 30, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    fill_ona   = PatternFill("solid", fgColor="EBF3FB")
    fill_ada   = PatternFill("solid", fgColor="EBF7EE")
    fill_att   = PatternFill("solid", fgColor="FFF2CC")
    fill_mil_v = PatternFill("solid", fgColor="E8F5E9")
    fill_mil_w = PatternFill("solid", fgColor="FCE4EC")

    # subject, category, subcategory, is_att, order, question, a,b,c,d, correct, qtype, parts, kw1, kw2
    examples = [
        # Ona tili — oddiy savollar
        ["onatili", "mavzu", "fonetika_tovushlar_tasnifi", "FALSE", None,
         "O'zbek tilida nechta unli tovush bor?", "5 ta", "6 ta", "7 ta", "8 ta", "B",
         "choice", 1, None, None],
        ["onatili", "mavzu", "morfologiya_m_ot", "FALSE", None,
         "Qaysi so'z ot turkumiga kiradi?", "yugurmoq", "baland", "maktab", "tez", "C",
         "choice", 1, None, None],
        ["onatili", "aralash", None, "FALSE", None,
         "Sintaksis nimani o'rganadi?", "So'z yasalishini", "Gap qurilishini", "Tovushlarni", "So'z ma'nosini", "B",
         "choice", 1, None, None],
        # Adabiyot — oddiy savollar
        ["adabiyot", "sinf", "7_2", "FALSE", None,
         "Navoiy qaysi asrda yashagan?", "XIV asr", "XV asr", "XVI asr", "XIII asr", "B",
         "choice", 1, None, None],
        ["adabiyot", "sheriy", None, "FALSE", None,
         "Tashbeh nima?", "O'xshatish", "Mubolag'a", "Takror", "Irsoli masal", "A",
         "choice", 1, None, None],
        # Attestatsiya — 1-bo'lim (subcategory="bolim_1")
        ["attestation", "attestation", "bolim_1", "TRUE", 1,
         "Fonetika nimani o'rganadi?", "So'z ma'nosini", "Tovush va harflarni", "Gap tuzilishini", "So'z yasalishini", "B",
         "choice", 1, None, None],
        ["attestation", "attestation", "bolim_1", "TRUE", 2,
         "Navoiy qaysi asrda yashagan?", "XIV asr", "XV asr", "XVI asr", "XIII asr", "B",
         "choice", 1, None, None],
        # Attestatsiya — 2-bo'lim (subcategory="bolim_2")
        ["attestation", "attestation", "bolim_2", "TRUE", 1,
         "Sintaksis nimani o'rganadi?", "So'z yasalishini", "Gap qurilishini", "Tovushlarni", "So'z ma'nosini", "B",
         "choice", 1, None, None],
        ["attestation", "attestation", "bolim_2", "TRUE", 2,
         "G'azal janrining asosiy belgisi nima?", "Bayt", "Qofiya", "Radif", "Maqta", "B",
         "choice", 1, None, None],
        # Milliy sertifikat — subject="milliy", variantli (1-35)
        ["milliy", "milliy", None, "TRUE", 1,
         "Fonetika nimani o'rganadi?", "So'z ma'nosini", "Tovush va harflarni", "Gap tuzilishini", "So'z yasalishini", "B",
         "choice", 1, None, None],
        ["milliy", "milliy", None, "TRUE", 2,
         "Navoiy qaysi asrda yashagan?", "XIV asr", "XV asr", "XVI asr", "XIII asr", "B",
         "choice", 1, None, None],
        # Milliy sertifikat — yozma 1 qism (36-38)
        ["milliy", "milliy", None, "TRUE", 36,
         "Fonetika fanining asosiy vazifasini izohlang.", None, None, None, None, None,
         "written", 1, "tovush, harflar, talaffuz", None],
        ["milliy", "milliy", None, "TRUE", 37,
         "Navoiy ijodining asosiy mavzusini tushuntiring.", None, None, None, None, None,
         "written", 1, "inson, muhabbat, ma'rifat", None],
        # Milliy sertifikat — yozma 2 qism (39-44)
        ["milliy", "milliy", None, "TRUE", 39,
         "Ot so'z turkumini ta'riflang va misol keltiring.", None, None, None, None, None,
         "written", 2, "ot, predmet, kim, nima", "misol, qo'shimcha"],
        ["milliy", "milliy", None, "TRUE", 40,
         "G'azal janrining xususiyatlarini izohlang.", None, None, None, None, None,
         "written", 2, "g'azal, bayt, radif, qofiya", "misol, Navoiy"],
    ]

    for i, row_data in enumerate(examples, start=2):
        subj     = row_data[0]
        is_yazma = len(row_data) > 11 and row_data[11] == "written"
        if subj == "milliy" and is_yazma:
            fill = fill_mil_w
        elif subj == "milliy":
            fill = fill_mil_v
        elif subj == "attestation":
            fill = fill_att
        elif subj == "onatili":
            fill = fill_ona
        else:
            fill = fill_ada
        for j, val in enumerate(row_data, start=1):
            cell = ws.cell(i, j)
            cell.value = val
            cell.fill  = fill

    # Qo'llanma
    ws2 = wb.create_sheet("Qo'llanma")
    ws2.column_dimensions['A'].width = 36
    ws2.column_dimensions['B'].width = 60

    hf2    = PatternFill("solid", fgColor="1F4E79")
    hfont2 = Font(bold=True, color="FFFFFF")
    blue   = Font(bold=True, color="1F4E79")
    red    = Font(bold=True, color="C00000")
    bold   = Font(bold=True)

    guide = [
        ("SAVOLLAR SHABLONI — ONA TILI VA ADABIYOT BOTI", ""),
        ("", ""),
        ("USTUN", "QIYMAT / IZOH"),
        ("subject",        "onatili | adabiyot | attestation | milliy"),
        ("category",       "mavzu | aralash | sinf | gazallar | sheriy | badiiy | attestation | milliy"),
        ("subcategory",    "Pastdagi jadvalga qarang"),
        ("is_attestation", "FALSE  (attestation/milliy uchun TRUE)"),
        ("order_num",      "Attestation: 1-35 | Milliy: 1-35 variantli, 36-38 yozma(1q), 39-44 yozma(2q)"),
        ("question",       "Savol matni"),
        ("a / b / c / d",  "Variant matni  (yozma savol uchun bo'sh qoldiring)"),
        ("correct",        "To'g'ri javob: A|B|C|D  (yozma savol uchun bo'sh)"),
        ("question_type",  "choice  |  written"),
        ("written_parts",  "1  yoki  2"),
        ("keywords_1",     "1-qism kalit so'zlari — vergul bilan: fonetika, tovush, unli"),
        ("keywords_2",     "2-qism kalit so'zlari — faqat written_parts=2 bo'lsa"),
        ("", ""),
        ("ESLATMA: difficulty ustuni yo'q!", ""),
        ("", ""),
        ("ATTESTATSIYA", "subject=attestation, category=attestation"),
        ("  subcategory", "bolim_1 | bolim_2 | bolim_3 | ... | bolim_10"),
        ("  Har bir bo'lim", "35 ta savol, variantli (A/B/C/D)"),
        ("  order_num", "Har bir bo'lim ichida 1 dan boshlanadi (1, 2, 3...)"),
        ("MILLIY SERTIFIKAT", "subject=milliy, category=milliy, fan ajratilmaydi"),
        ("  1-35 savol",  "question_type=choice"),
        ("  36-38 savol", "question_type=written, written_parts=1"),
        ("  39-44 savol", "question_type=written, written_parts=2"),
        ("", ""),
        ("ONA TILI — SUBCATEGORY", ""),
        ("category = mavzu", ""),
        ("fonetika",                    "Fonetika (barcha)"),
        ("fonetika_tovushlar_tasnifi",  "Tovushlar tasnifi"),
        ("fonetika_tovush_ozgarishi",   "Tovush o'zgarishlari"),
        ("imlo",                        "Imlo (barcha)"),
        ("imlo_togri_yozilgan",         "Qaysi so'z to'g'ri yozilgan"),
        ("imlo_imloviy_xato",           "Gapda imloviy xato"),
        ("morfemika",                   "Morfemika (barcha)"),
        ("morfemika_qoshimchalar",      "Qo'shimchalar tasnifi"),
        ("morfemika_tub_yasama",        "Tub va yasama so'zlar"),
        ("leksikologiya",               "Leksikologiya (barcha)"),
        ("leksikologiya_oz_kochma",     "O'z va ko'chma ma'no"),
        ("leksikologiya_omonimlik",     "Omonimlik"),
        ("leksikologiya_paronimlik",    "Paronimlik"),
        ("leksikologiya_ibora",         "Ibora va tasviriy ifodalar"),
        ("leksikologiya_lugatlar",      "Lug'atlardan"),
        ("morfologiya_m",               "Morfologiya mustaqil (barcha)"),
        ("morfologiya_m_ot",            "Ot"), ("morfologiya_m_sifat", "Sifat"),
        ("morfologiya_m_son",           "Son"), ("morfologiya_m_olmosh", "Olmosh"),
        ("morfologiya_m_ravish",        "Ravish"), ("morfologiya_m_fel", "Fe'l"),
        ("morfologiya_y",               "Morfologiya yordamchi (barcha)"),
        ("morfologiya_y_boglovchilar",  "Bog'lovchilar"),
        ("morfologiya_y_komakchilar",   "Ko'makchilar"),
        ("morfologiya_y_yuklamalar",    "Yuklamalar"),
        ("morfologiya_a",               "Alohida so'z turkumlari"),
        ("sintaksis",                   "Sintaksis (barcha)"),
        ("sintaksis_sozlar_boglashi",   "So'zlarning bog'lanishi"),
        ("sintaksis_gap_bolaklari",     "Gap bo'laklari"),
        ("sintaksis_qoshma_gaplar",     "Qo'shma gaplar"),
        ("matnlar",                     "Matnlar (barcha)"),
        ("matnlar_ilmiy_matn",          "Ilmiy matnlar"),
        ("matnlar_badiiy_matn",         "Badiiy matnlar"),
        ("punktuatsiya",                "Punktuatsiya"),
        ("uslubiyat",                   "Uslubiyat (barcha)"),
        ("uslubiyat_qoshimchalar_uslubiyat", "Qo'shimchalar uslubiyati"),
        ("uslubiyat_sozlar_uslubiyat",  "So'zlar uslubiyati"),
        ("category = aralash",          "subcategory = bo'sh"),
        ("", ""),
        ("ADABIYOT — SUBCATEGORY", ""),
        ("category = sinf", ""),
        ("5 / 5_1 / 5_2 / 5_3 / 5_4",  "5-sinf barcha / bob bo'yicha"),
        ("6...10 uchun ham xuddi shunday", ""),
        ("11 / 11_1 / 11_2 / 11_3 / 11_4", "11-sinf"),
        ("ATTESTATSIYA — SUBCATEGORY", ""),
        ("bolim_1",  "1-bo'lim savollar (order_num=1,2,3...)"),
        ("bolim_2",  "2-bo'lim savollar"),
        ("bolim_3",  "3-bo'lim savollar"),
        ("bolim_4",  "4-bo'lim savollar"),
        ("bolim_5",  "5-bo'lim savollar"),
        ("bolim_6",  "6-bo'lim savollar"),
        ("bolim_7",  "7-bo'lim savollar"),
        ("bolim_8",  "8-bo'lim savollar"),
        ("bolim_9",  "9-bo'lim savollar"),
        ("bolim_10", "10-bo'lim savollar"),
        ("", ""),
        ("category = gazallar",  "subcategory = bo'sh"),
        ("category = sheriy",    "subcategory = bo'sh"),
        ("category = badiiy",    "subcategory = bo'sh"),
        ("category = aralash",   "subcategory = bo'sh"),
    ]

    for i, (col_a, col_b) in enumerate(guide, start=1):
        ca = ws2.cell(i, 1, col_a)
        cb = ws2.cell(i, 2, col_b)
        if i == 1:
            ca.font = Font(bold=True, size=13, color="1F4E79")
        elif col_a == "USTUN":
            ca.fill = hf2; ca.font = hfont2
            cb.fill = hf2; cb.font = hfont2
        elif col_a in ("ONA TILI — SUBCATEGORY", "ADABIYOT — SUBCATEGORY",
                       "ATTESTATSIYA", "MILLIY SERTIFIKAT"):
            ca.font = blue
        elif "ESLATMA" in col_a:
            ca.font = red
        elif col_a.startswith("category ="):
            ca.font = bold
        ca.alignment = Alignment(wrap_text=True, vertical='top')
        cb.alignment = Alignment(wrap_text=True, vertical='top')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    await message.answer_document(
        document=BufferedInputFile(buf.read(), filename="savollar_shablon.xlsx"),
        caption=(
            "📋 <b>Excel shablon</b>\n\n"
            "🔵 Ona tili savollar\n"
            "🟢 Adabiyot savollar\n"
            "🟡 Attestatsiya (1-35 variantli, fan ajratilmaydi)\n"
            "🟩 Milliy sertifikat variantli (1-35)\n"
            "🩷 Milliy sertifikat yozma (36-44)\n\n"
            "1️⃣ Yuklab oling  2️⃣ To'ldiring  3️⃣ Yuboring\n"
            "📌 <b>Qo'llanma</b> varaqasini o'qing!"
        ),
        parse_mode="HTML"
    )
@router.message(F.document)
async def excel_import_upload(message: Message):
    if not is_admin(message): return

    doc = message.document
    if not doc.file_name or not doc.file_name.endswith('.xlsx'):
        return  # xlsx emas — e'tiborsiz

    await message.answer("⏳ Fayl o'qilmoqda...")

    # Faylni yuklab olish
    from aiogram import Bot
    bot: Bot = message.bot
    file = await bot.get_file(doc.file_id)
    buf = io.BytesIO()
    await bot.download_file(file.file_path, buf)
    buf.seek(0)

    wb = openpyxl.load_workbook(buf)
    ws = wb.active

    VALID_SUBJECTS   = {'onatili', 'adabiyot', 'attestation', 'milliy'}
    VALID_CATEGORIES = {'mavzu', 'aralash', 'sinf', 'gazallar',
                        'sheriy', 'badiiy', 'attestation', 'milliy'}
    VALID_CORRECT    = {'A', 'B', 'C', 'D'}
    # Attestatsiya subcategory — bolim_1 ... bolim_10 yoki None
    VALID_ATTEST_SUBCATS = {f"bolim_{i}" for i in range(1, 11)} | {None, ""}

    added   = 0
    skipped = 0
    errors  = []

    # Format aniqlash — sarlavhaga qarab
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    cols = [str(c or '').strip().lower() for c in (first_row or [])]
    has_difficulty   = 'difficulty' in cols
    has_written_cols = 'question_type' in cols

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not any(row):
            continue

        try:
            if has_difficulty:
                # Eski format (12 ustun)
                subject        = str(row[0] or '').strip().lower()
                category       = str(row[1] or '').strip().lower()
                subcategory    = str(row[2] or '').strip() or None
                is_attestation = str(row[4] or '').strip().upper() == 'TRUE'
                order_num      = int(row[5]) if row[5] else None
                question_text  = str(row[6] or '').strip()
                option_a       = str(row[7] or '').strip() or None
                option_b       = str(row[8] or '').strip() or None
                option_c       = str(row[9] or '').strip() or None
                option_d       = str(row[10] or '').strip() or None
                correct        = str(row[11] or '').strip().upper() or None
                question_type  = 'choice'
                written_parts  = 1
                keywords_1     = None
                keywords_2     = None
            elif has_written_cols:
                # Yangi format (15 ustun)
                subject        = str(row[0] or '').strip().lower()
                category       = str(row[1] or '').strip().lower()
                subcategory    = str(row[2] or '').strip() or None
                is_attestation = str(row[3] or '').strip().upper() == 'TRUE'
                order_num      = int(row[4]) if row[4] else None
                question_text  = str(row[5] or '').strip('\r\n \t') if row[5] else ''
                option_a       = str(row[6] or '').strip('\r\n \t') or None
                option_b       = str(row[7] or '').strip('\r\n \t') or None
                option_c       = str(row[8] or '').strip('\r\n \t') or None
                option_d       = str(row[9] or '').strip('\r\n \t') or None
                correct        = str(row[10] or '').strip().upper() or None
                question_type  = str(row[11] or 'choice').strip().lower() or 'choice'
                written_parts  = int(row[12]) if row[12] else 1
                keywords_1     = str(row[13] or '').strip() or None
                keywords_2     = str(row[14] or '').strip() or None
            else:
                # O'rta format (11 ustun)
                subject        = str(row[0] or '').strip().lower()
                category       = str(row[1] or '').strip().lower()
                subcategory    = str(row[2] or '').strip() or None
                is_attestation = str(row[3] or '').strip().upper() == 'TRUE'
                order_num      = int(row[4]) if row[4] else None
                question_text  = str(row[5] or '').strip('\r\n \t') if row[5] else ''
                option_a       = str(row[6] or '').strip('\r\n \t') or None
                option_b       = str(row[7] or '').strip('\r\n \t') or None
                option_c       = str(row[8] or '').strip('\r\n \t') or None
                option_d       = str(row[9] or '').strip('\r\n \t') or None
                correct        = str(row[10] or '').strip().upper() or None
                question_type  = 'choice'
                written_parts  = 1
                keywords_1     = None
                keywords_2     = None

            if subject not in VALID_SUBJECTS:
                errors.append(f"Qator {row_num}: subject '{subject}' noto'g'ri")
                skipped += 1; continue
            if category not in VALID_CATEGORIES:
                errors.append(f"Qator {row_num}: category '{category}' noto'g'ri")
                skipped += 1; continue
            if not question_text:
                errors.append(f"Qator {row_num}: savol matni bo'sh")
                skipped += 1; continue
            if question_type == 'choice':
                if correct not in VALID_CORRECT:
                    errors.append(f"Qator {row_num}: correct '{correct}' noto'g'ri")
                    skipped += 1; continue
                if not all([option_a, option_b, option_c, option_d]):
                    errors.append(f"Qator {row_num}: variantlar to'liq emas")
                    skipped += 1; continue

            await add_question(
                subject=subject, category=category,
                subcategory=subcategory, difficulty=None,
                is_attestation=is_attestation, order_num=order_num,
                question_text=question_text,
                option_a=option_a, option_b=option_b,
                option_c=option_c, option_d=option_d,
                correct_answer=correct,
                question_type=question_type,
                written_parts=written_parts,
                keywords_1=keywords_1,
                keywords_2=keywords_2,
            )
            added += 1

        except Exception as e:
            errors.append(f"Qator {row_num}: {e}")
            skipped += 1

    # Natija
    text = (
        f"✅ <b>Import tugadi!</b>\n\n"
        f"✅ Qo'shildi: <b>{added} ta</b>\n"
        f"❌ O'tkazildi: <b>{skipped} ta</b>\n"
    )
    if errors:
        error_text = "\n".join(errors[:10])
        if len(errors) > 10:
            error_text += f"\n... va yana {len(errors)-10} ta xato"
        text += f"\n⚠️ <b>Xatolar:</b>\n<code>{error_text}</code>"

    await message.answer(text, parse_mode="HTML", reply_markup=admin_keyboard())