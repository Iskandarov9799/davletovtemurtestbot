"""
Excel import testlari.
To'g'ri va noto'g'ri fayllarni tekshiradi.
"""
import io
import pytest
import openpyxl
from database.db import count_questions, add_question


def make_xlsx(rows: list) -> io.BytesIO:
    """Test uchun xlsx fayl yaratish."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "subject", "category", "subcategory", "is_attestation",
        "order_num", "question", "a", "b", "c", "d", "correct",
        "question_type", "written_parts", "keywords_1", "keywords_2"
    ]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_xlsx(buf: io.BytesIO) -> tuple:
    """
    Excel faylini parse qilish — admin.py dagi logikani qayta ishlatish.
    (added, skipped, errors) qaytaradi.
    """
    import asyncio
    from database.db import add_question

    wb  = openpyxl.load_workbook(buf)
    ws  = wb.active

    VALID_SUBJECTS   = {'onatili', 'adabiyot', 'attestation', 'milliy'}
    VALID_CATEGORIES = {'mavzu', 'aralash', 'sinf', 'gazallar',
                        'sheriy', 'badiiy', 'attestation', 'milliy'}
    VALID_CORRECT    = {'A', 'B', 'C', 'D'}

    added = skipped = 0
    errors = []

    async def _run():
        nonlocal added, skipped
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
            try:
                subject        = str(row[0] or '').strip().lower()
                category       = str(row[1] or '').strip().lower()
                subcategory    = str(row[2] or '').strip() or None
                is_attestation = str(row[3] or '').strip().upper() == 'TRUE'
                order_num      = int(row[4]) if row[4] else None
                question_text  = str(row[5] or '').strip()
                option_a       = str(row[6] or '').strip() or None
                option_b       = str(row[7] or '').strip() or None
                option_c       = str(row[8] or '').strip() or None
                option_d       = str(row[9] or '').strip() or None
                correct        = str(row[10] or '').strip().upper() or None
                question_type  = str(row[11] or 'choice').strip().lower() or 'choice'
                written_parts  = int(row[12]) if row[12] else 1
                keywords_1     = str(row[13] or '').strip() or None
                keywords_2     = str(row[14] or '').strip() or None

                if subject not in VALID_SUBJECTS:
                    errors.append(f"Qator {row_num}: subject '{subject}' noto'g'ri")
                    skipped += 1; continue
                if category not in VALID_CATEGORIES:
                    errors.append(f"Qator {row_num}: category '{category}' noto'g'ri")
                    skipped += 1; continue
                if not question_text:
                    errors.append(f"Qator {row_num}: savol matni bo'sh")
                    skipped += 1; continue
                if question_type == 'choice':
                    if correct not in VALID_CORRECT:
                        errors.append(f"Qator {row_num}: correct '{correct}' noto'g'ri")
                        skipped += 1; continue
                    if not all([option_a, option_b, option_c, option_d]):
                        errors.append(f"Qator {row_num}: variantlar to'liq emas")
                        skipped += 1; continue

                await add_question(
                    subject=subject, category=category,
                    subcategory=subcategory, difficulty=None,
                    is_attestation=is_attestation, order_num=order_num,
                    question_text=question_text,
                    option_a=option_a, option_b=option_b,
                    option_c=option_c, option_d=option_d,
                    correct_answer=correct,
                    question_type=question_type,
                    written_parts=written_parts,
                    keywords_1=keywords_1, keywords_2=keywords_2,
                )
                added += 1
            except Exception as e:
                errors.append(f"Qator {row_num}: {e}")
                skipped += 1

    asyncio.get_event_loop().run_until_complete(_run())
    return added, skipped, errors


# ════════════════════════════════════════════════
# EXCEL IMPORT TESTLARI
# ════════════════════════════════════════════════

class TestExcelImport:

    @pytest.mark.asyncio
    async def test_valid_choice_questions(self, test_db):
        """To'g'ri variantli savollarni import qilish."""
        rows = [
            ["onatili", "mavzu", "fonetika", "FALSE", None,
             "O'zbek tilida nechta unli bor?",
             "5 ta", "6 ta", "7 ta", "8 ta", "B", "choice", 1, None, None],
            ["adabiyot", "aralash", None, "FALSE", None,
             "Navoiy qaysi asrda yashagan?",
             "XIV", "XV", "XVI", "XIII", "B", "choice", 1, None, None],
        ]
        buf = make_xlsx(rows)
        added, skipped, errors = parse_xlsx(buf)
        assert added == 2
        assert skipped == 0
        assert errors == []

    @pytest.mark.asyncio
    async def test_written_question_import(self, test_db):
        """Yozma savollarni import qilish."""
        rows = [
            ["milliy", "milliy", None, "TRUE", 36,
             "Fonetikani izohlang.", None, None, None, None, None,
             "written", 1, "tovush, harf", None],
            ["milliy", "milliy", None, "TRUE", 39,
             "Ot so'z turkumini ta'riflang.", None, None, None, None, None,
             "written", 2, "ot, predmet", "misol, qo'shimcha"],
        ]
        buf = make_xlsx(rows)
        added, skipped, errors = parse_xlsx(buf)
        assert added == 2
        assert skipped == 0

    @pytest.mark.asyncio
    async def test_invalid_subject(self, test_db):
        """Noto'g'ri subject — o'tkazib yuboriladi."""
        rows = [
            ["fizika", "mavzu", None, "FALSE", None,
             "Savol matni", "A", "B", "C", "D", "A", "choice", 1, None, None],
        ]
        buf = make_xlsx(rows)
        added, skipped, errors = parse_xlsx(buf)
        assert added == 0
        assert skipped == 1
        assert "subject" in errors[0]

    @pytest.mark.asyncio
    async def test_invalid_correct_answer(self, test_db):
        """Noto'g'ri to'g'ri javob — o'tkazib yuboriladi."""
        rows = [
            ["onatili", "aralash", None, "FALSE", None,
             "Savol matni", "A", "B", "C", "D", "E", "choice", 1, None, None],
        ]
        buf = make_xlsx(rows)
        added, skipped, errors = parse_xlsx(buf)
        assert added == 0
        assert skipped == 1

    @pytest.mark.asyncio
    async def test_empty_question_text(self, test_db):
        """Bo'sh savol matni — o'tkazib yuboriladi."""
        rows = [
            ["onatili", "aralash", None, "FALSE", None,
             "", "A", "B", "C", "D", "A", "choice", 1, None, None],
        ]
        buf = make_xlsx(rows)
        added, skipped, errors = parse_xlsx(buf)
        assert added == 0
        assert skipped == 1

    @pytest.mark.asyncio
    async def test_missing_options(self, test_db):
        """Variantlar to'liq emas — o'tkazib yuboriladi."""
        rows = [
            ["onatili", "aralash", None, "FALSE", None,
             "Savol matni", "A", None, "C", "D", "A", "choice", 1, None, None],
        ]
        buf = make_xlsx(rows)
        added, skipped, errors = parse_xlsx(buf)
        assert added == 0
        assert skipped == 1

    @pytest.mark.asyncio
    async def test_mixed_valid_invalid(self, test_db):
        """Aralash — ba'zilar o'tadi, ba'zilari o'tkazib yuboriladi."""
        rows = [
            ["onatili", "aralash", None, "FALSE", None,
             "To'g'ri savol", "A", "B", "C", "D", "A", "choice", 1, None, None],
            ["noto'g'ri_subject", "aralash", None, "FALSE", None,
             "Xato savol", "A", "B", "C", "D", "A", "choice", 1, None, None],
            ["adabiyot", "sheriy", None, "FALSE", None,
             "Yana to'g'ri", "A", "B", "C", "D", "B", "choice", 1, None, None],
        ]
        buf = make_xlsx(rows)
        added, skipped, errors = parse_xlsx(buf)
        assert added == 2
        assert skipped == 1

    @pytest.mark.asyncio
    async def test_empty_rows_ignored(self, test_db):
        """Bo'sh qatorlar e'tiborga olinmaydi."""
        rows = [
            ["onatili", "aralash", None, "FALSE", None,
             "Savol", "A", "B", "C", "D", "A", "choice", 1, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
        ]
        buf = make_xlsx(rows)
        added, skipped, errors = parse_xlsx(buf)
        assert added == 1

    @pytest.mark.asyncio
    async def test_attestation_questions(self, test_db):
        """Attestatsiya savollarini import qilish."""
        rows = [
            ["attestation", "attestation", None, "TRUE", 1,
             "Fonetika savoli", "A", "B", "C", "D", "B", "choice", 1, None, None],
            ["attestation", "attestation", None, "TRUE", 2,
             "Morfologiya savoli", "A", "B", "C", "D", "C", "choice", 1, None, None],
        ]
        buf = make_xlsx(rows)
        added, skipped, errors = parse_xlsx(buf)
        assert added == 2
        cnt = await count_questions(subject="attestation", is_attestation=True)
        assert cnt == 2